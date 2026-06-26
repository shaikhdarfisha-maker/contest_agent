from datetime import datetime
from modules.tracker import ContestTracker
from modules.utils import derive_attempt_windows
from openpyxl import load_workbook

t = ContestTracker()

w1 = derive_attempt_windows(datetime(2027,1,1,21,0), datetime(2027,1,10,21,0))
r1 = t.append_contest(module='Advanced DSA 4',
                      batch_name='Advanced DSA 4: NV Contest January 2027',
                      windows=w1, use_concatenate_formula=True)

w2 = derive_attempt_windows(datetime(2027,2,1,21,0), datetime(2027,2,10,21,0))
r2 = t.append_contest(module='Advanced DSA 4',
                      batch_name='Advanced DSA 4: NV Contest February 2027',
                      windows=w2, use_concatenate_formula=False)

wb = load_workbook('data/NV_contests_Tracker_Q2-2026_.xlsx')
ws = wb['Academy New Contests ']
print()
print('=== Column B (Batch name) comparison ===')
print('Row', r1, '(Option A, formula) :', ws.cell(r1, 2).value)
print('Row', r2, '(Option B, plain)   :', ws.cell(r2, 2).value)

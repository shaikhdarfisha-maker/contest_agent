"""
tracker.py
==========
Appends a single contest row to the production NV Contest Tracker while
treating that workbook as the source of truth.

Hard rules (per ops constraints):
  * Do NOT alter the schema: no new columns, no renamed headers.
  * Preserve existing formulas, formatting, validations and column order.
  * Write ONLY the columns the ops team fills manually
    (Module, Batch Name, and the 4 attempt start/end datetimes).
  * Columns K-N (No. of Attempts / Final Deadline / Days Remaining /
    Contest Status) are spreadsheet formulas and are deliberately left for
    Excel to compute - we never write literals there.
  * All internal IDs / timestamps / errors go to the SQLite store, not here.

Append logic is inferred from the workbook itself: we find the first empty
row beneath the existing data block in the configured sheet and write into it,
mirroring how existing rows are laid out (including, optionally, the
CONCATENATE batch-name formula style the sheet already uses).
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    TRACKER_COLS,
    TRACKER_FIRST_DATA_ROW,
    TRACKER_SHEET,
    TRACKER_WORKBOOK,
)
from modules.logger import get_logger
from modules.utils import AttemptWindow, DuplicateContestError, TrackerUpdateError

log = get_logger(__name__)


class ContestTracker:
    """Safe, schema-preserving appender for the NV Contest Tracker."""

    def __init__(
        self,
        workbook_path: Path = TRACKER_WORKBOOK,
        sheet_name: str = TRACKER_SHEET,
    ) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Tracker workbook not found: {workbook_path}")

    # ---------------------------------------------------------------------- #
    def _backup(self) -> Path:
        """Copy the workbook before mutating it, so a bad write is recoverable."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = self.workbook_path.with_name(
            f"{self.workbook_path.stem}.bak_{ts}{self.workbook_path.suffix}"
        )
        shutil.copy2(self.workbook_path, backup)
        log.debug("Tracker backed up to %s", backup.name)
        return backup

    @staticmethod
    def _first_empty_row(ws: Worksheet, module_col: int, start_row: int) -> int:
        """
        Infer the append point: the first row at/after start_row whose module
        cell (col A) is empty. Mirrors how a human appends to the bottom of the
        existing block rather than trusting ws.max_row (which can over-report).
        """
        row = start_row
        while ws.cell(row=row, column=module_col).value not in (None, ""):
            row += 1
        return row

    def _propagate_formula_columns(self, ws: Worksheet, target_row: int) -> None:
        """
        Replicate the K-N formula columns from a reference data row into
        target_row, re-pointing every relative row reference to target_row.

        We read the formula pattern from the sheet itself (the nearest filled
        row above target_row) rather than hardcoding it, so if ops change the
        formulas the agent follows along automatically.
        """
        import re

        formula_cols = (11, 12, 13, 14)  # K, L, M, N
        # Find a reference row above target_row that has a formula in col K.
        ref_row: Optional[int] = None
        for r in range(target_row - 1, TRACKER_FIRST_DATA_ROW - 1, -1):
            val = ws.cell(row=r, column=formula_cols[0]).value
            if isinstance(val, str) and val.startswith("="):
                ref_row = r
                break
        if ref_row is None:
            log.debug("No reference formula row found; leaving K-N blank.")
            return

        # Re-point row numbers in the formula. Matches a column letter followed
        # by the reference row number as a whole token (e.g. C5, J5, L5).
        token = re.compile(rf"(?<=[A-Za-z]){ref_row}(?![0-9])")
        for col in formula_cols:
            ref_val = ws.cell(row=ref_row, column=col).value
            if isinstance(ref_val, str) and ref_val.startswith("="):
                new_formula = token.sub(str(target_row), ref_val)
                ws.cell(row=target_row, column=col, value=new_formula)
        log.debug(
            "Propagated K-N formulas from row %d to row %d", ref_row, target_row
        )

    def _row_exists(self, ws: Worksheet, batch_name: str) -> bool:
        """
        Duplicate detection. The Batch Name column often holds a CONCATENATE
        formula rather than a literal, so openpyxl returns the formula text, not
        the evaluated name. We therefore match on either (a) a literal batch
        name in column B, or (b) a row whose module (col A) + the formula's
        suffix reconstruct to the same name.
        """
        b_col = TRACKER_COLS["batch_name"]
        a_col = TRACKER_COLS["module"]
        target = batch_name.strip().lower()
        for row in range(TRACKER_FIRST_DATA_ROW, ws.max_row + 1):
            val = ws.cell(row=row, column=b_col).value
            if not isinstance(val, str):
                continue
            if val.strip().startswith("="):
                # Reconstruct "<module>: <suffix>" from A + the formula suffix.
                module_val = ws.cell(row=row, column=a_col).value
                suffix = ""
                if '"' in val:
                    parts = val.split('"')
                    # e.g. =CONCATENATE(A5,": NV Contest June 2026") -> parts[1]
                    if len(parts) >= 2:
                        suffix = parts[1].lstrip(": ").strip()
                if module_val:
                    reconstructed = f"{str(module_val).strip()}: {suffix}".lower()
                    if reconstructed == target:
                        return True
            elif val.strip().lower() == target:
                return True
        return False

    # ---------------------------------------------------------------------- #
    def append_contest(
        self,
        *,
        module: str,
        batch_name: str,
        windows: list[AttemptWindow],
        use_concatenate_formula: bool = False,  # default: plain-text batch name
        propagate_formulas: bool = True,
        dry_run: bool = False,
    ) -> int:
        """
        Append one contest row. Returns the 1-based row index written.

        Only the manual columns are populated. Re-attempts beyond those provided
        are left blank (the sheet tolerates 1-4 filled attempts).
        """
        if len(windows) < 1:
            raise TrackerUpdateError("At least the main contest window is required.")

        # Open WITHOUT data_only so existing formulas are preserved on save.
        wb = load_workbook(self.workbook_path)
        if self.sheet_name not in wb.sheetnames:
            raise TrackerUpdateError(
                f"Sheet '{self.sheet_name}' not in tracker. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[self.sheet_name]

        if self._row_exists(ws, batch_name):
            raise DuplicateContestError(
                f"A row for batch '{batch_name}' already exists in "
                f"'{self.sheet_name}'."
            )

        row = self._first_empty_row(
            ws, TRACKER_COLS["module"], TRACKER_FIRST_DATA_ROW
        )

        # Column A: module name (exactly as ops enter it).
        ws.cell(row=row, column=TRACKER_COLS["module"], value=module)

        # Column B: batch name. The sheet's existing rows use a CONCATENATE
        # formula referencing column A; we mirror that style by default so the
        # row is indistinguishable from a hand-entered one, but allow a plain
        # literal as a fallback.
        if use_concatenate_formula:
            suffix = batch_name.split(":", 1)[1].strip() if ":" in batch_name else ""
            formula = f'=CONCATENATE(A{row},": {suffix}")'
            ws.cell(row=row, column=TRACKER_COLS["batch_name"], value=formula)
        else:
            ws.cell(row=row, column=TRACKER_COLS["batch_name"], value=batch_name)

        # Columns C-J: up to 4 attempt windows as real datetimes (so the
        # sheet's MAX/COUNT formulas in K-N evaluate correctly).
        attempt_cols = [
            ("a1_start", "a1_end"),
            ("a2_start", "a2_end"),
            ("a3_start", "a3_end"),
            ("a4_start", "a4_end"),
        ]
        for win, (start_key, end_key) in zip(windows, attempt_cols):
            ws.cell(row=row, column=TRACKER_COLS[start_key], value=win.start)
            ws.cell(row=row, column=TRACKER_COLS[end_key], value=win.end)

        # Columns K-N are formula columns in the sheet (No. of Attempts /
        # Final Deadline / Days Remaining / Contest Status). We do NOT invent
        # values; instead we replicate the existing formula pattern, re-pointed
        # at this row, so the appended row computes exactly like the others.
        if propagate_formulas:
            self._propagate_formula_columns(ws, row)

        if dry_run:
            log.info("[dry-run] Would append '%s' at row %d", batch_name, row)
            return row

        self._backup()
        wb.save(self.workbook_path)
        log.info("Tracker updated: appended '%s' at row %d", batch_name, row)
        return row

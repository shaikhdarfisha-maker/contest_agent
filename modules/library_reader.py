"""
library_reader.py
=================
Reads Library__All_Programs.xlsx and resolves (program, module) -> library.

The workbook has one sheet per program with slightly different column layouts
(see config.PROGRAMS). The operator always supplies the program, which removes
the cross-program ambiguity. Within a program a module can still map to more
than one library (e.g. Academy "Advance Programming Concepts" has Java and
Python variants); we prefer a "Live"/non-deprecated row when the sheet exposes
a status column, otherwise we raise AmbiguousLibraryError with the candidates
so the caller can decide.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from config import LIBRARY_WORKBOOK, PROGRAMS, SheetSpec
from modules.logger import get_logger
from modules.utils import AmbiguousLibraryError, LibraryNotFoundError

log = get_logger(__name__)

# Tokens in a library name that indicate a deprecated/non-live row, used as a
# fallback when the sheet has no explicit status column.
_DEPRECATED_TOKENS = ("(na)", "(old", "inverted", "oldv", "(2024)", "(old)")


@dataclass(frozen=True)
class LibraryMatch:
    """Resolved library for a module within a program."""

    module: str
    program: str
    library_name: str
    library_link: Optional[str]
    library_id: Optional[str]
    num_attempts: int = 4  # per-module attempt count (from "Num Attempts" column)
    skill_eval_label: Optional[str] = None  # override for the Mandatory Skill
    # Evaluation checkbox match, when the module name doesn't textually match
    # the checkbox label Scaler shows (e.g. module "Linux Shell Scripting and
    # Computer Systems 2" vs. checkbox "Linux Certification Contest").


def _norm(text: object) -> str:
    return str(text or "").strip().lower()


def _extract_library_id(link: Optional[str]) -> Optional[str]:
    """Pull the trailing numeric id out of .../edit-library/277 style links."""
    if not link:
        return None
    tail = str(link).rstrip("/").rsplit("/", 1)[-1]
    return tail if tail.isdigit() else None


class LibraryReader:
    """Loads and queries a single library workbook."""

    def __init__(self, workbook_path: Path = LIBRARY_WORKBOOK) -> None:
        self.workbook_path = workbook_path
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Library workbook not found: {workbook_path}")
        # read_only + data_only: we only need the resolved values, fast.
        self._wb = load_workbook(
            self.workbook_path, read_only=True, data_only=True
        )

    # ---------------------------------------------------------------------- #
    def _header_index(self, rows: list[tuple], spec: SheetSpec) -> dict[str, int]:
        """Find column indices by matching header text (case-insensitive)."""
        header = rows[0]
        index: dict[str, int] = {}
        wanted = {
            "module": _norm(spec.module_col),
            "library": _norm(spec.library_col),
            "link": _norm(spec.link_col) if spec.link_col else None,
            "status": _norm(spec.status_col) if spec.status_col else None,
            "attempts": _norm(spec.attempts_col) if spec.attempts_col else None,
            "skill_eval": _norm(spec.skill_eval_col) if spec.skill_eval_col else None,
        }
        for col_i, cell in enumerate(header):
            cell_norm = _norm(cell)
            for key, want in wanted.items():
                if want and cell_norm == want:
                    index[key] = col_i
        if "module" not in index or "library" not in index:
            raise LibraryNotFoundError(
                f"Could not locate module/library headers in sheet "
                f"'{spec.sheet_name}'. Found headers: {header}"
            )
        return index

    def _candidates(self, program: str, module: str) -> list[LibraryMatch]:
        spec = PROGRAMS.get(program.lower())
        if spec is None:
            raise LibraryNotFoundError(
                f"Unknown program '{program}'. Known: {list(PROGRAMS)}"
            )
        if spec.sheet_name not in self._wb.sheetnames:
            raise LibraryNotFoundError(
                f"Sheet '{spec.sheet_name}' not found in {self.workbook_path.name}"
            )

        ws = self._wb[spec.sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise LibraryNotFoundError(f"Sheet '{spec.sheet_name}' is empty")

        idx = self._header_index(rows, spec)
        target = _norm(module)
        matches: list[LibraryMatch] = []

        for row in rows[1:]:
            if _norm(row[idx["module"]]) != target:
                continue
            link = row[idx["link"]] if "link" in idx else None
            _attempts_raw = row[idx["attempts"]] if "attempts" in idx else None
            try:
                _num_attempts = int(_attempts_raw) if _attempts_raw is not None and str(_attempts_raw).strip() else 4
            except (ValueError, TypeError):
                _num_attempts = 4
            _skill_eval_raw = row[idx["skill_eval"]] if "skill_eval" in idx else None
            _skill_eval_label = (
                str(_skill_eval_raw).strip() if _skill_eval_raw and str(_skill_eval_raw).strip() else None
            )
            matches.append(
                LibraryMatch(
                    module=str(row[idx["module"]]).strip(),
                    program=program.lower(),
                    library_name=str(row[idx["library"]]).strip(),
                    library_link=str(link).strip() if link else None,
                    library_id=_extract_library_id(link),
                    num_attempts=_num_attempts,
                    skill_eval_label=_skill_eval_label,
                )
            )
        return matches

    # ---------------------------------------------------------------------- #
    def resolve(
        self, program: str, module: str, *, prefer_live: bool = True
    ) -> LibraryMatch:
        """
        Resolve a module to a single library within a program.

        Raises:
            LibraryNotFoundError   - no row for that module.
            AmbiguousLibraryError  - several plausible rows and none preferred.
        """
        matches = self._candidates(program, module)
        if not matches:
            raise LibraryNotFoundError(
                f"No library found for module '{module}' in program '{program}'."
            )
        if len(matches) == 1:
            log.info("Library resolved: %s -> %s", module, matches[0].library_name)
            return matches[0]

        # Multiple candidates: filter to non-deprecated first, then take the
        # first row. Row order in the workbook IS the priority order — the
        # preferred library (e.g. NV Contests) is always prepended first.
        if prefer_live:
            live = [
                m
                for m in matches
                if not any(tok in _norm(m.library_name) for tok in _DEPRECATED_TOKENS)
            ]
            if live:
                matches = live

        log.info(
            "Library resolved (first of %d): %s -> %s",
            len(matches),
            module,
            matches[0].library_name,
        )
        return matches[0]

    def resolve_by_name_only(self, program: str, library_name: str) -> LibraryMatch:
        """Find any row in the program sheet whose library name matches, ignoring module."""
        spec = PROGRAMS.get(program.lower())
        if spec is None:
            raise LibraryNotFoundError(f"Unknown program '{program}'.")
        if spec.sheet_name not in self._wb.sheetnames:
            raise LibraryNotFoundError(
                f"Sheet '{spec.sheet_name}' not found in {self.workbook_path.name}"
            )
        ws = self._wb[spec.sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise LibraryNotFoundError(f"Sheet '{spec.sheet_name}' is empty")
        idx = self._header_index(rows, spec)
        target = _norm(library_name)
        for row in rows[1:]:
            if _norm(row[idx["library"]]) == target:
                link = row[idx["link"]] if "link" in idx else None
                return LibraryMatch(
                    module=str(row[idx["module"]]).strip(),
                    program=program.lower(),
                    library_name=str(row[idx["library"]]).strip(),
                    library_link=str(link).strip() if link else None,
                    library_id=_extract_library_id(link),
                )
        raise LibraryNotFoundError(
            f"Fallback library '{library_name}' not found in program '{program}'."
        )

    def all_module_names(self, program: str) -> list[str]:
        """Return all unique module names for a program sheet, sorted."""
        spec = PROGRAMS.get(program.lower())
        if spec is None or spec.sheet_name not in self._wb.sheetnames:
            return []
        ws = self._wb[spec.sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return []
        try:
            idx = self._header_index(rows, spec)
        except LibraryNotFoundError:
            return []
        names: set[str] = set()
        for row in rows[1:]:
            mod = str(row[idx["module"]] or "").strip()
            if mod:
                names.add(mod)
        return sorted(names, key=str.casefold)

    def all_library_names(self) -> list[str]:
        """Return all unique library names across every program sheet, sorted."""
        names: set[str] = set()
        for spec in PROGRAMS.values():
            if spec.sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[spec.sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            try:
                idx = self._header_index(rows, spec)
            except LibraryNotFoundError:
                continue
            for row in rows[1:]:
                lib = str(row[idx["library"]] or "").strip()
                if lib:
                    names.add(lib)
        return sorted(names, key=str.casefold)

    def resolve_explicit(
        self, program: str, module: str, library_name: str
    ) -> LibraryMatch:
        """Resolve when the operator has already named the exact library."""
        target = _norm(library_name)
        for m in self._candidates(program, module):
            if _norm(m.library_name) == target:
                return m
        raise LibraryNotFoundError(
            f"Library '{library_name}' not found for module '{module}' "
            f"in program '{program}'."
        )
